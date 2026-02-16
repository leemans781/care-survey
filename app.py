# -*- coding: utf-8 -*-
"""
Created on Tue Jan 13 13:12:48 2026

@author: SashaLeemans
"""

# app.py — Streamlit AHP survey met integratie van de code van care 0.005 
import os
import streamlit as st
import numpy as np
import pandas as pd
from numpy.linalg import eigvals
from openpyxl import Workbook
from io import BytesIO
import json
import matplotlib.pyplot as plt


# BESTANDEN EN INSTELLINGEN
st.set_page_config(page_title="AHP Survey", layout="wide")

RESP_DIR = "responses"  # map waarin CSV's van deelnemers komen
os.makedirs(RESP_DIR, exist_ok=True)

ADMIN_CODE = "secret123"
CONFIG_FILE = "config.json"

# Dit moet ervoor zorgen dat de respondenten pas kunnen invullen, als de beheerder de criteria heeft ingevuld
# Er valt namelijk nog niks in te vullen, als de criteria nog niet bekend zijn
# Dit zelfde pas ik doe voor de alternatieven. 
if "criteria" not in st.session_state:
    st.session_state.criteria = []
if "criteria_locked" not in st.session_state:
    st.session_state.criteria_locked = False
       
if "alternatives" not in st.session_state:
    st.session_state.alternatives = []
if "alternatives_locked" not in st.session_state:
    st.session_state.alternatives_locked = False
    
if "participant_name" not in st.session_state:
    st.session_state.participant_name = ""
if "participant_id" not in st.session_state:
    st.session_state.participant_id = ""
if "criteria_submitted" not in st.session_state:
    st.session_state.criteria_submitted = False    
    
# Config laden (gedeeld voor alle gebruikers)
if os.path.exists(CONFIG_FILE):
    with open(CONFIG_FILE, "r") as f:
        config = json.load(f)

    st.session_state.criteria = config.get("criteria", [])
    st.session_state.alternatives = config.get("alternatives", [])
    st.session_state.survey_open = config.get("survey_open", False)
    #st.session_state.criteria_locked = config.get("survey_open", False)
    #st.session_state.alternatives_locked = config.get("survey_open", False)


def weights_colmean(M: np.ndarray) -> np.ndarray:
    """Bereken AHP-gewichten via kolomnormalisatie + rijgemiddelden (som=1)."""
    colsum = M.sum(axis=0)
    norm = M / colsum
    w = norm.mean(axis=1)
    return w / w.sum()

def saaty_cr(M: np.ndarray, w: np.ndarray) -> float:
    """Saaty Consistency Ratio (voor n<=10)."""
    lam = np.mean((M @ w) / w)
    n = M.shape[0]
    ri_map = {1:0.0, 2:0.0, 3:0.58, 4:0.90, 5:1.12, 6:1.24, 7:1.32, 8:1.41, 9:1.45, 10:1.49}
    ri = ri_map.get(n, None)
    if ri is None:
        # Voor n>10: advies is Alonso-Lamata
        return np.nan
    ci = (lam - n) / (n - 1)
    return 0.0 if ri == 0 else ci / ri

def alo_cr(M: np.ndarray) -> float:
    """Alonso-Lamata CR (voor n>10)."""
    n = M.shape[0]
    lam = max(np.real(eigvals(M)))
    return (lam - n) / (2.7699 * n - 4.3513 - n)

def consolidate_matrices(matrices: list[np.ndarray]) -> np.ndarray:
    """AIJ: element-wise product + k-de wortel (geometrisch gemiddelde) tot consolidated matrix."""
    if not matrices:
        return None
    P = np.ones_like(matrices[0], dtype=float)
    for A in matrices:
        P *= A
    k = len(matrices)
    G = P ** (1.0 / k)
    return G


def calculate_homogeneity(priorities: list[dict]) -> float:
    """
    Bereken homogeniteit op basis van Shannon-entropie.
    priorities = lijst van dicts met gewichten per respondent.
    """
    n = len(priorities)  # aantal respondenten
    cats = list(priorities[0].keys())  # categorieën
    catCnt = len(cats)

    # Gemiddelde distributie (groep)
    avg = {c: 0.0 for c in cats}
    for p in priorities:
        for c in cats:
            avg[c] += p[c]
    for c in cats:
        avg[c] /= n

    # Entropie van groep (gamma)
    gamma = -sum(avg[c] * np.log(avg[c]) for c in cats if avg[c] > 0)

    # Entropie van individuen (alpha)
    alpha_sum = 0.0
    for p in priorities:
        alpha_sum += -sum(p[c] * np.log(p[c]) for c in cats if p[c] > 0)
    alpha = alpha_sum / n

    # Beta = verschil
    beta = gamma - alpha
    sim = (1. / np.exp(beta) - 1. / catCnt) / (1. - 1. / catCnt)  # schaal zoals in jouw collega’s code
    return sim

def cosine_similarity_matrix(X: np.ndarray) -> np.ndarray:
    """
    Bereken cosine similarity matrix zonder sklearn.
    X: shape (n_respondenten, n_criteria)
    """
    # Normaliseer elke vector
    norms = np.linalg.norm(X, axis=1, keepdims=True)
    X_norm = X / norms

    # Cosine similarity = dot product van genormaliseerde vectors
    return X_norm @ X_norm.T


def calculate_consensus(priorities: list[dict]) -> float:
    """
    Bereken consensus via gemiddelde cosine similarity (zonder sklearn).
    """
    cats = list(priorities[0].keys())
    vectors = np.array([[p[c] for c in cats] for p in priorities])

    cosine_sim_matrix = cosine_similarity_matrix(vectors)

    # Gemiddelde off-diagonal similarity
    total = 0.0
    count = 0
    for i in range(cosine_sim_matrix.shape[0]):
        for j in range(cosine_sim_matrix.shape[1]):
            if i != j:
                total += cosine_sim_matrix[i, j]
                count += 1

    return total / count if count > 0 else 0.0


def safe_float(x):
    try:
        return float(x)
    except:
        return np.nan


# Sidebar: moduskeuze
st.sidebar.header("Navigatie")
mode = st.sidebar.radio("Kies modus", ["Deelnemer (invullen)", "Admin (groepresultaat)"])


# BASISINFORMATIE 
criteria = st.session_state.criteria
n = len(criteria)

alternatieven = st.session_state.alternatives
a = len(alternatieven)

st.header("Analytic Hierarchy Process (AHP) - Beslissingsanalyse")
st.write("### Bij dutch process innovators (dpi)")
st.write("Vergelijk criteria en alternatieven om tot een gewogen beslissing te komen")

# DEELNEMER MODUS
if mode == "Deelnemer (invullen)":
    
    # Survey pas openen als criteria & alternatieven bekend zijn
    if not st.session_state.get('survey_open', False):
        st.warning("De survey is nog niet geopend.")
        st.stop()
        
    st.subheader("Start jouw beoordeling")

    # Invullen naam/e-mail
    participant_name = st.text_input("Jouw naam (of e-mail)", value=st.session_state.participant_name, placeholder="Naam of e-mail")
    st.session_state.participant_name = participant_name.strip()
    if not st.session_state.participant_name:
        st.info("Vul je naam/e-mail in om verder te gaan.")
        st.stop()
        
    #if not st.session_state.participant_id:
    if "participant_id" not in st.session_state:     
        st.session_state.participant_id = "".join(ch for ch in st.session_state.participant_name if ch.isalnum() or ch in ("_", "-", ".")).lower()

    st.write(f"Aantal criteria: **{n}**")
    st.write(f"Aantal alternatieven:  **{a}**")
    st.markdown("---")    
    

    tabs = st.tabs(["Criteria", "Alternatieven"])
    # Criteria invullen
    with tabs[0]:
        st.subheader("Criteria vergelijken")
        st.caption("Kies per paar welk criterium belangrijker is en hoe sterk.")
     
        vals = {}
        for i in range(n):
            for j in range(i + 1, n):
                key = f"{criteria[i]} vs {criteria[j]}"

                with st.container(border=True):
                    st.write(f"**Vergelijk:** {criteria[i]} vs {criteria[j]}")

                    col1, col2 = st.columns([4, 2])
                    with col1:
                        side = st.radio(
                            "Keuze",
                            [
                                f"{criteria[i]} > {criteria[j]}",
                                f"{criteria[i]} ≈ {criteria[j]}",
                                f"{criteria[j]} > {criteria[i]}",
                            ],
                            index=1,  # standaard op gelijk, voelt neutraler
                            horizontal=True,
                            key=key + "_side",
                            label_visibility="collapsed",
                        )
                    with col2:
                        mag = st.slider("Sterkte (1 = zwak, 9 = zeer sterk)", 1, 9, 3, key=key + "_mag")

                    # Waarde bepalen
                    if "≈" in side:
                        v = 1.0
                    elif side.startswith(criteria[i]):  # i > j
                        v = float(mag)
                    else:  # j > i
                        v = 1.0 / float(mag)

                    vals[(i, j)] = v
    
        # Volledige matrix opbouwen
        A = np.ones((n, n), dtype=float)
        for (i, j), v in vals.items():
            A[i, j] = v
            A[j, i] = 1.0 / v
    
        st.write("**Jouw pairwise matrix**")
        st.dataframe(pd.DataFrame(A, columns=criteria, index=criteria))
    
        # Gewichten + CR
        w = weights_colmean(A)
        cr = saaty_cr(A, w) if n <= 10 else alo_cr(A)
    
        st.subheader("Jouw prioriteiten")
        st.write(pd.DataFrame({"Criteria": criteria, "Weight (%)": (w * 100).round(2)}))
        st.metric("Consistency Ratio (CR)", f"{cr * 100:.1f}%")
    
        # Opslaan
        st.markdown("---")
        if st.button("Verstuur en opslaan"):
            pid = st.session_state.participant_id
            st.session_state.criteria_submitted = True
            # Bestandsnaam veilig maken
            safe_name = "".join(ch for ch in participant_name if ch.isalnum() or ch in ("_", "-", "."))
            out_path = os.path.join(RESP_DIR, f"{safe_name}.csv")
            pd.DataFrame(A, columns=criteria, index=criteria).to_csv(out_path, index=True)
            #st.success(f"Inzending opgeslagen: `{out_path}`")
            st.info("Je kunt het tabblad sluiten. Bedankt voor het invullen!")
     
    # Alternatieven invullen    
    with tabs[1]:
        st.subheader("Alternatieven vergelijken")
        st.caption("Vergelijk de alternatieven per criterium.")
        
        if not st.session_state.criteria_submitted:
            st.warning("Je moet eerst de criteria afronden en versturen voordat je alternatieven kunt waarderen.")
            st.stop()
            
        # Kleine status bovenaan
        st.info(f"Ingelogd als: **{st.session_state.participant_name}**  |  Criteria afgerond: ✅")
        
        alternatives = st.session_state.alternatives
        criteria = st.session_state.criteria
        all_alt_matrices = {}
        

        # Loop over elk criterium
        for crit in criteria:
            with st.expander(f" Criterium: {crit}", expanded=False):
                st.markdown("Kies per paar welk alternatief beter scoort op dit criterium, en hoe sterk.")
            
            n_alt = len(alternatives)
            vals = {}

            # Boven-diagonaal invoer van alternatieven
            for i in range(n_alt):
                for j in range(i+1, n_alt):
                    key = f"{crit}_{alternatives[i]}_vs_{alternatives[j]}"
                    with st.container():
                        col1, col2 = st.columns([3,2])
                        with col1:
                            side = st.radio(
                                key,
                                [
                                    f"{alternatives[i]} > {alternatives[j]}",
                                    f"{alternatives[i]} ≈ {alternatives[j]}",
                                    f"{alternatives[j]} > {alternatives[i]}"
                                ],
                                index=0,
                                horizontal=True
                            )
                        with col2:
                            mag = st.slider("Sterkte (1 = zwak, 9 = zeer sterk)", 1, 9, 3, key=key + "_mag")
                    
                    # Bereken waarde
                    if "≈" in side:
                        v = 1.0
                    elif side.startswith(alternatives[i]):
                        v = float(mag)
                    else:
                        v = 1.0 / float(mag)
                    
                    vals[(i,j)] = v

            # Bouw volledige matrix
            A = np.ones((n_alt, n_alt), dtype=float)
            for (i,j), v in vals.items():
                A[i,j] = v
                A[j,i] = 1.0 / v
            
            all_alt_matrices[crit] = A.copy()

            st.write("**Jouw pairwise matrix voor dit criterium**")
            st.dataframe(pd.DataFrame(A, columns=alternatives, index=alternatives))

            # Gewichten + CR
            w = weights_colmean(A)
            cr = saaty_cr(A, w) if n_alt <= 10 else alo_cr(A)

            st.subheader("Prioriteiten voor dit criterium")
            st.write(pd.DataFrame({"Alternatief": alternatives, "Weight (%)": (w*100).round(2)}))
            st.metric("Consistency Ratio (CR)", f"{cr*100:.1f}%")

        # Opslaan van alle alternatieven-matrices per criterium
        st.markdown("---")
        if st.button("Verstuur en opslaan alternatieven"):
            safe_name = "".join(ch for ch in participant_name if ch.isalnum() or ch in ("_", "-", "."))
            alt_dir = os.path.join(RESP_DIR, "alternatives")
            os.makedirs(alt_dir, exist_ok=True)

            for crit, Acrit in all_alt_matrices.items():
                # Hier zou je dezelfde matrices A moeten opslaan per criterium
                safe_crit = "".join(ch for ch in crit if ch.isalnum() or ch in ("_", "-"))
                out_path = os.path.join(alt_dir, f"{safe_name}_{safe_crit}.csv")
                pd.DataFrame(Acrit, columns=alternatives, index=alternatives).to_csv(out_path, index=True)
            st.success(f"Inzendingen voor alle criteria opgeslagen in '{alt_dir}'")
            st.info("Bedankt voor het invullen van de alternatieven!")


# ADMIN MODUS
else:
    st.subheader("Admin — groepsresultaat (alleen voor jou)")

    code = st.text_input("Admin code", type="password")
    if code != ADMIN_CODE:
        st.warning("Voer de juiste admin code in om groepsresultaat te zien.")
        st.stop()
        
    # Criteria instellen    
    st.markdown("### Criteria instellen (alleen admin)")
    criteria_input = st.text_area("Voer criteria in (één per regel)", value="\n".join(st.session_state.criteria))
    
    if st.button("Criteria opslaan"):
        new_criteria = [c.strip() for c in criteria_input.splitlines() if c.strip()]
        if len(new_criteria) < 2:
            st.error("Minimaal 2 criteria vereist.")
        else:
            st.session_state.criteria = new_criteria
            #st.session_state.criteria_locked = True
            st.success("Criteria opgeslagen.")    
            
    # De volgende stap: alternatieven introduceren. 
    st.markdown("### Alternatieven instellen (alleen admin)")
    alternatives_input = st.text_area("Voer alternatieven in (één per regel)", value="\n".join(st.session_state.alternatives))
    
    if st.button("Alternatieven opslaan"):
        new_alternatives = [a.strip() for a in alternatives_input.splitlines() if a.strip()]
        if len(new_alternatives) < 2:
            st.error("Minimaal 2 alternatieven vereist.")
        else:
            st.session_state.alternatives = new_alternatives
            st.success("Alternatieven opgeslagen.")
     
    # Survey open / gesloten status         
    if (len(st.session_state.criteria) >= 2 and len(st.session_state.alternatives) >= 2):
        st.session_state.criteria_locked = True
        st.session_state.alternatives_locked = True
        
    st.markdown("### Survey status")
    can_open = len(st.session_state.criteria) >= 2 and len(st.session_state.alternatives) >= 2

    if not can_open:
        st.info("Voer minimaal 2 criteria en 2 alternatieven in om de survey te kunnen openen.")
    
    # Alleen tonen en laten klikken als voldoende gegevens
    if can_open:
        if st.button("Open survey voor deelnemers"):
            config = {
                "criteria": st.session_state.criteria,
                "alternatives": st.session_state.alternatives,
                "survey_open": True
            }
            with open(CONFIG_FILE, "w") as f:
                json.dump(config, f)
            st.success("Survey is nu geopend voor alle deelnemers.")
            
    # De volgende stap is het tonen van de resultaten. 
    # Dit is opgedeeld in de resultaten van de criteria en de alternatieven
    st.markdown("### Resultaten")
    tabs = st.tabs(["Criteria", "Alternatieven"])
    with tabs[0]:
        st.subheader("Criteria — groepsresultaten")
        # Lees alle responses die overeenkomen met deze criteria-dimensie (n x n)
        files = [f for f in os.listdir(RESP_DIR) if f.endswith(".csv")]
        st.write(f"Gevonden inzendingen: **{len(files)}**")
        matrices = []
        bad_files = []
        for f in files:
            path = os.path.join(RESP_DIR, f)
            try:
                df = pd.read_csv(path, index_col=0)
                # controle: zelfde dimensie en zelfde criteria set?
                if df.shape != (n, n):
                    bad_files.append(f"{f} (dimensie {df.shape}, verwacht {(n,n)})")
                    continue
                # optioneel: check index/kolomnamen gelijk aan criteria
                # Bij afwijkende namen kun je alleen op dimensie matchen in MVP
                A = df.values.astype(float)
                matrices.append(A)
            except Exception as e:
                bad_files.append(f"{f} (error: {e})")
    
        if bad_files:
            st.warning("Sommige files konden niet gebruikt worden:\n- " + "\n- ".join(bad_files))
    
        if not matrices:
            st.info("Geen bruikbare responses gevonden in de map 'responses/'.")
            st.stop()
    
        # Consolidated decision matrix via AIJ
        G = consolidate_matrices(matrices)
        st.write("**Consolidated Decision Matrix (groep)**")
        st.dataframe(pd.DataFrame(G, columns=criteria, index=criteria))
    
        # Groepsgewichten + CR
        wg = weights_colmean(G)
        group_cr = saaty_cr(G, wg) if n <= 10 else alo_cr(G)
    
        # Overzichtstabel
        st.subheader("Groepsprioriteiten")
        df_grp = pd.DataFrame({
            "Criteria": criteria,
            "Weight (%)": (wg * 100).round(2)
        })
        # Rang bepalen
        df_grp["Rank"] = df_grp["Weight (%)"].rank(ascending=False, method="dense").astype(int)
        st.write(df_grp)
        
        # Visualisatie (bar plot)
        fig, ax = plt.subplots(figsize=(4,2), dpi=80)
        criteria_names = df_grp["Criteria"]
        weights = df_grp["Weight (%)"]
        bars = ax.bar(criteria_names, weights)
        ax.set_xlabel("Criteria", fontsize=8)
        ax.set_ylabel("Gewicht (%)", fontsize=8)
        ax.set_ylim(0,100)
        ax.tick_params(axis='both', labelsize=8)
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2, height, f"{height:.2f}%", ha="center", va="bottom", fontsize=8)
            
        st.pyplot(fig, use_container_width=False)
        #st.metric("Group Consistency Ratio (CR)", f"{group_cr * 100:.1f}%")
        
        # Bereken homogeniteit en consensus
        priorities_list = []
        for f in files:
            df = pd.read_csv(os.path.join(RESP_DIR, f), index_col=0)
            weights = weights_colmean(df.values)
            priorities_list.append({criteria[i]: weights[i] for i in range(len(criteria))})
        
        homogeneity = calculate_homogeneity(priorities_list)
        consensus = calculate_consensus(priorities_list)
        
        st.subheader("Groepshomogeniteit & Consensus")
        st.metric("Homogeniteit (S)", f"{homogeneity:.3f}")
        st.metric("Consensus (S*)", f"{consensus:.3f}")
        
        # Interpretatie
        def interpret(value):
            if value >= 0.8:
                return "Hoog"
            elif value >= 0.6:
                return "Matig"
            else:
                return "Laag"
        
        st.write(f"Interpretatie homogeniteit: {interpret(homogeneity)}")
        st.progress(homogeneity)
        st.write(f"Interpretatie consensus: {interpret(consensus)}")
        st.progress(consensus)
    
        # Export knoppen
        st.markdown("---")
        colA, colB, colC = st.columns(3)
        with colA:
            # consolidated matrix download
            cm_bytes = pd.DataFrame(G, columns=criteria, index=criteria).to_csv(index=True).encode("utf-8")
            st.download_button("Download consolidated matrix (CSV)", cm_bytes, file_name="consolidated_matrix.csv")
        with colB:
            # group weights download
            gw_bytes = df_grp.to_csv(index=False).encode("utf-8")
            st.download_button("Download group weights (CSV)", gw_bytes, file_name="group_weights.csv")
        with colC:
            # log van gebruikte files
            log_bytes = "\n".join(files).encode("utf-8")
            st.download_button("Download list of participant files (TXT)", log_bytes, file_name="participants_used.txt")
            
        st.markdown("---")

        # Export knop, die de individuele resultaten opslaat in Excel. 
        wb = Workbook()
        wb.remove(wb.active)  # verwijder standaard sheet
        
        for f in files:
            path = os.path.join(RESP_DIR, f)
            df = pd.read_csv(path, index_col=0)
            participant = os.path.splitext(f)[0]
            
            ws = wb.create_sheet(title=participant[:31])  # Excel max 31 chars
            
            A = df.values.astype(float)
            weights = weights_colmean(A)
            cr = saaty_cr(A, weights) if n <= 10 else alo_cr(A)
            
            # ---- Matrix schrijven ----
            ws.append(["Pairwise Matrix"])
            ws.append([])
            
            ws.append([""] + list(df.columns))
            for idx, row in df.iterrows():
                ws.append([idx] + list(row.values))
            
            ws.append([])
            ws.append([])
            
            # ---- Prioriteiten schrijven ----
            ws.append(["Prioriteiten"])
            ws.append(["Criteria", "Weight (%)"])
            
            for i in range(n):
                ws.append([criteria[i], round(weights[i] * 100, 2)])
            
            ws.append([])
            ws.append(["Consistency Ratio", round(cr * 100, 2)])
        
        # In-memory opslaan
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        st.download_button(label="Download individuele resultaten (Excel)", data=output,file_name="individual_results.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        st.caption("MVP: responses staan lokaal in de map 'responses/'. In Streamlit Cloud blijven ze bewaard zolang de app niet opnieuw wordt gedeployed. Voor productie: gebruik een database of Blob Storage.")
    
    with tabs[1]:
        st.subheader("Alternatieven — groepsresultaten")
        
        #Lees alle responses van alternatieven
        alt_dir = os.path.join(RESP_DIR, "alternatives")
        if not os.path.exists(alt_dir):
            st.info("Nog geen alternatieven beoordeeld door deelnemers.")
            st.stop()
            
        files = [f for f in os.listdir(alt_dir) if f.endswith(".csv")]
        # st.write(f"Gevonden inzendingen: **{len(files)}**")
        
        participants = set(f.split("_")[0] for f in files)
        st.write(f"Gevonden deelnemers: **{len(participants)}**")

        
        if not files:
            st.info("Geen alternatieven-responses gevonden.")
            st.stop()
            
        # Voor elke criterium, consolideer matrix
        consolidated_per_crit = {}
        bad_files = []
        
        for crit in criteria:
            matrices = []
            for f in files:
                if crit in f:
                    path = os.path.join(alt_dir, f)
                    try:
                        df = pd.read_csv(path, index_col=0)
                        if df.shape != (len(alternatieven), len(alternatieven)):
                            bad_files.append(f"{f} (dimensie {df.shape}, verwacht {(len(alternatieven), len(alternatieven))}")
                            continue
                        matrices.append(df.values.astype(float))
                    except Exception as e:
                        bad_files.append(f"{f} (error: {e})")
            if matrices:
                consolidated_per_crit[crit] = consolidate_matrices(matrices)
        
        if bad_files:
            st.warning("Sommige bestanden konden niet gebruikt worden:\n- " + "\n- ".join(bad_files))
            
        if not consolidated_per_crit:
            st.info("Geen bruikbare alternatieven gevonden.")
            st.stop()
            
        # Toon een tabel per criterium
        st.markdown("### Geconsolideerde gewichten per criterium")
        for crit, mat in consolidated_per_crit.items():
            w = weights_colmean(mat)
            st.write(f"**Criterium: {crit}**")
            df_w = pd.DataFrame({"Alternatief": alternatieven, "Gewicht (%)": (w*100).round(2)})
            df_w["Rang"] = df_w["Gewicht (%)"].rank(ascending=False, method='dense').astype(int)
            st.dataframe(df_w)
        
        criteria_weights = wg
    
        total_scores = np.zeros(len(alternatieven))
        for i, crit in enumerate(criteria):
            w_alt = weights_colmean(consolidated_per_crit[crit])
            total_scores += w_alt * wg[i]
    
        df_total = pd.DataFrame({
            "Alternatief": alternatieven,
            "Totale gewogen score (%)": (total_scores*100).round(2)
        })
        df_total["Rang"] = df_total["Totale gewogen score (%)"].rank(ascending=False, method="dense").astype(int)
    
        st.markdown("### Totale gewogen scores over alle criteria")
        st.dataframe(df_total)
        
        best_alt = df_total.loc[df_total["Totale gewogen score (%)"].idxmax(), "Alternatief"]
        st.markdown(f'### Beste optie alternatief: {best_alt}')
        fig, ax = plt.subplots(figsize=(5,3))
        alternatives_names = df_total["Alternatief"]
        scores = df_total["Totale gewogen score (%)"]
        bars = ax.bar(alternatives_names, scores, score='skyblue')
        ax.set_ylabel("Totale gewogen score (%)")
        ax.set_ylim(0,100)
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2, height, f"{height:.1f}%", ha="center", va="bottom", fontsize=8)
            
        plt.tight_layout()
        st.pyplot(fig)
        # Downloadknoppen
        st.markdown("---")
        colA, colB, colC, colD = st.columns(4)
        with colA:
            csv_bytes = pd.concat([pd.DataFrame(w) for w in consolidated_per_crit.values()], axis=1).to_csv(index=False).encode("utf-8")
            st.download_button("Download overzicht per criterium", csv_bytes, "alternatives_per_criterion.csv")
        with colB:
            csv_bytes = df_total.to_csv(index=False).encode("utf-8")
            st.download_button("Download totale gewichten", csv_bytes, "total_alternative_weights.csv")
        with colC:
            log_bytes = "\n".join(files).encode("utf-8")
            st.download_button("Download lijst deelnemers", log_bytes, "participants_used.txt")
        with colD:
            # Individuele resultaten in Excel
            wb = Workbook()
            wb.remove(wb.active)
            for f in files:
                path = os.path.join(alt_dir, f)
                df = pd.read_csv(path, index_col=0)
                participant = os.path.splitext(f)[0][:31]
                ws = wb.create_sheet(title=participant)
                ws.append(["Pairwise Matrix"])
                ws.append([])
                ws.append([""] + list(df.columns))
                for idx, row in df.iterrows():
                    ws.append([idx] + list(row.values))
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            st.download_button("Download individuele resultaten (Excel)", data=output,
                               file_name="individual_alternative_results.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                
            